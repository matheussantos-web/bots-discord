import discord
from discord.ext import commands
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# Importa as variáveis do seu arquivo de configuração
from config import colecao_pontos, MONGO_URI, CARGOS, CARGOS_PERMITIDOS_ADICIONAR, CARGOS_PERMITIDOS_REMOVER, LIMITE_ADICAO_POR_VEZ

class Economia(commands.Cog):
    def __init__(self, bot):
        self.bot = bot

    @commands.command(name="pontos", aliases=["saldo"])
    async def ver_pontos(self, ctx, membro: discord.Member = None):
        membro_alvo = membro or ctx.author 
        
        if not MONGO_URI:
            return await ctx.send("⚠️ O Banco de Dados de pontos está inacessível no momento.")

        documento = await colecao_pontos.find_one({"_id": str(membro_alvo.id)})
        saldo = documento.get("pontos", 0) if documento else 0
        
        embed = discord.Embed(
            title="💰 Banco da Guilda - Saldo de Atividade",
            description=f"O jogador {membro_alvo.mention} possui **{saldo} pontos** acumulados.",
            color=discord.Color.green()
        )
        embed.set_thumbnail(url=membro_alvo.display_avatar.url)
        embed.set_footer(text="Participe de conteúdos em calls com 5+ membros para pontuar!")
        await ctx.send(embed=embed)


    @commands.command(name="adicionarpontos")
    async def adicionar_pontos_manual(self, ctx, membro: discord.Member, quantidade: int):
        ids_permitidos = [CARGOS.get(nome) for nome in CARGOS_PERMITIDOS_ADICIONAR if CARGOS.get(nome)]
        tem_permissao = any(cargo.id in ids_permitidos for cargo in ctx.author.roles)
        
        if not tem_permissao and not ctx.author.guild_permissions.administrator:
            return await ctx.send("❌ Acesso Negado: Seu cargo não tem permissão para adicionar pontos.")

        if not MONGO_URI:
            return await ctx.send("⚠️ Banco de dados offline.")
            
        if quantidade <= 0:
            return await ctx.send("❌ A quantidade deve ser um número positivo.")
            
        if quantidade > LIMITE_ADICAO_POR_VEZ:
            return await ctx.send(f"❌ Medida de Segurança: Você só pode adicionar até **{LIMITE_ADICAO_POR_VEZ} pontos** por vez.")

        await colecao_pontos.update_one(
            {"_id": str(membro.id)},
            {"$inc": {"pontos": quantidade}},
            upsert=True
        )
        await ctx.send(f"✅ Foram adicionados **{quantidade} pontos** na carteira de {membro.mention} com sucesso!")


    @commands.command(name="removerpontos")
    async def remover_pontos_manual(self, ctx, membro: discord.Member, quantidade: int):
        ids_permitidos = [CARGOS.get(nome) for nome in CARGOS_PERMITIDOS_REMOVER if CARGOS.get(nome)]
        tem_permissao = any(cargo.id in ids_permitidos for cargo in ctx.author.roles)
        
        if not tem_permissao and not ctx.author.guild_permissions.administrator:
            return await ctx.send("❌ Acesso Restrito: Apenas a Alta Cúpula pode remover pontos dos membros.")

        if not MONGO_URI:
            return await ctx.send("⚠️ Banco de dados offline.")
            
        if quantidade <= 0:
            return await ctx.send("❌ A quantidade deve ser um número positivo.")

        documento = await colecao_pontos.find_one({"_id": str(membro.id)})
        saldo_atual = documento.get("pontos", 0) if documento else 0

        if saldo_atual < quantidade:
            return await ctx.send(f"❌ Operação cancelada. O membro possui apenas **{saldo_atual} pontos**.")

        await colecao_pontos.update_one(
            {"_id": str(membro.id)},
            {"$inc": {"pontos": -quantidade}},
            upsert=True
        )
        await ctx.send(f"🛑 Foram removidos **{quantidade} pontos** da carteira de {membro.mention}.")


    @commands.command(name="relatorio")
    async def gerar_relatorio_pontos(self, ctx):
        ids_permitidos = [CARGOS.get(nome) for nome in CARGOS_PERMITIDOS_REMOVER if CARGOS.get(nome)]
        tem_permissao = any(cargo.id in ids_permitidos for cargo in ctx.author.roles)
        
        if not tem_permissao and not ctx.author.guild_permissions.administrator:
            return await ctx.send("❌ Acesso Restrito: Apenas a Alta Cúpula pode gerar o relatório geral de atividade.")

        if not MONGO_URI:
            return await ctx.send("⚠️ Banco de dados offline.")

        mensagem_aviso = await ctx.send("🔄 **Extraindo dados da nuvem e montando a planilha...** Aguarde um instante.")

        try:
            cursor = colecao_pontos.find({})
            documentos = await cursor.to_list(length=None)

            if not documentos:
                return await mensagem_aviso.edit(content="❌ O banco de dados está vazio no momento.")

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Relatório da Guilda"

            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            zebra_fill = PatternFill(start_color="F2F6FA", end_color="F2F6FA", fill_type="solid")
            font_header = Font(color="FFFFFF", bold=True)
            align_center = Alignment(horizontal="center", vertical="center")

            headers = ["ID do Jogador", "Nick na Guilda (Discord)", "Pontos Acumulados"]
            ws.append(headers)

            for col_num, _ in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = font_header
                cell.fill = header_fill
                cell.alignment = align_center

            linha_atual = 2
            for doc in documentos:
                id_discord = int(doc["_id"])
                pontos = doc.get("pontos", 0)

                membro = ctx.guild.get_member(id_discord)
                nome_jogador = membro.display_name if membro else "Membro Desligado"

                ws.append([str(id_discord), nome_jogador, pontos])

                if linha_atual % 2 == 0:
                    for col_idx in range(1, 4):
                        ws.cell(row=linha_atual, column=col_idx).fill = zebra_fill
                
                linha_atual += 1

            ws.column_dimensions["A"].width = 25
            ws.column_dimensions["B"].width = 40
            ws.column_dimensions["C"].width = 20

            ws.freeze_panes = "A2"

            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            nome_arquivo = f"Relatorio_Pontos_{datetime.now().strftime('%d_%m_%Y')}.xlsx"
            arquivo_discord = discord.File(fp=buffer, filename=nome_arquivo)
            
            await mensagem_aviso.delete()
            await ctx.send(content="📊 **Relatório Extraído com Sucesso!**\nAqui está a planilha oficial de atividade da guilda atualizada agora:", file=arquivo_discord)

        except Exception as e:
            await mensagem_aviso.edit(content=f"⚠️ Erro ao gerar a planilha: {e}")

# Função obrigatória para o main.py carregar este arquivo
async def setup(bot):
    await bot.add_cog(Economia(bot))