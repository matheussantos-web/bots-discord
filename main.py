import discord
from discord.ext import commands, tasks
import aiohttp
import os
from keep_alive import keep_alive
import asyncio
from datetime import datetime, timedelta, timezone
import motor.motor_asyncio
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ==========================================
#  CONFIGURAÇÕES INICIAIS E INTENTS
# ==========================================

intents = discord.Intents.default()
intents.message_content = True
intents.members = True
intents.voice_states = True
intents.presences = True

# O help_command=None desativa o comando de ajuda feio padrão do Discord
bot = commands.Bot(command_prefix="!", intents=intents, help_command=None)

# ==========================================
#  VARIÁVEIS DO SEU SERVIDOR
# ==========================================

GUILDA_ALBION_ID = "oZercpSURfeSz9_7Mpg1-w"  
ALIANCA_ALBION_ID = "mCKNk3EyQ8qLeNMhozFIZg" 
CANAIS_GERADORES_IDS = [1519180718406635570, 1486876393035010148] 
CANAL_LIMPEZA_ID = 1519195690503245835

MENSAGEM_CLASSES_ID = 1519507112156463195  

REACOES_CLASSES = {
    "🛡️": 1519187254042693652,  
    "💖": 1519187260774551661,  
    "🏹": 1519187263056122017,  
    "⚔️": 1519187266000523285,  
}

TAG_GUILDA = "[DH]" 
TAG_ALIANCA = "[ALLY]"

CARGOS = {
    "aliado": 1519187015307104276,
    "friends": 1519186964853555210, 
    "peleco": 1519187046051086497,  
    "DIE HARD": 1519186777838190592, 
    "recrutador": 1519186606836154428,
    "moderador": 1519186575139799061,
    "caller novato" : 1519186854614663208,
    "caller": 1519186818392920064,
    "SUB-LIDER": 1519186609310793728,
    "lider": 1519186611315933345,   
}   

# ==========================================
#  CONFIGURAÇÕES DO SISTEMA DE PONTOS (MONGODB)
# ==========================================
MINIMO_PESSOAS_CALL = 5
PONTOS_POR_CICLO = 1
MULTIPLICADOR_CALLER = 5

LIMITE_ADICAO_POR_VEZ = 10 

CARGOS_PERMITIDOS_ADICIONAR = ["lider", "SUB-LIDER", "caller", "caller novato"] 
CARGOS_PERMITIDOS_REMOVER = ["lider", "SUB-LIDER"]

MONGO_URI = os.getenv('MONGO_URI')
if MONGO_URI:
    mongo_client = motor.motor_asyncio.AsyncIOMotorClient(MONGO_URI)
    db = mongo_client["guilda_bot"]  
    colecao_pontos = db["pontos"]    
else:
    print("⚠️ MONGO_URI não encontrada! O sistema de pontos está desativado.")

# ==========================================
# SISTEMA DE LIMPEZA AUTOMÁTICA
# ==========================================

@tasks.loop(minutes=30)
async def limpeza_automatica():
    if CANAL_LIMPEZA_ID == 0:
        return 

    canal = bot.get_channel(CANAL_LIMPEZA_ID)
    if not canal:
        return

    limite_tempo = datetime.now(timezone.utc) - timedelta(hours=6)

    def verificar_mensagem(msg):
        return not msg.pinned 

    try:
        deletadas = await canal.purge(limit=None, before=limite_tempo, check=verificar_mensagem)
        if len(deletadas) > 0:
            print(f"🧹 Faxina concluída! {len(deletadas)} mensagens apagadas no canal de limpeza.")
    except Exception as e:
        print(f"⚠️ Erro ao tentar limpar o chat: {e}")

# ==========================================
# SISTEMA DE AUDITORIA DE MEMBROS
# ==========================================

@tasks.loop(hours=24)
async def auditoria_guilda():
    print("🔍 Iniciando ronda de auditoria diária na API do Albion...")
    
    guilda_discord = bot.guilds[0] 
    
    cargos_gerenciados = []
    for id_cargo in CARGOS.values():
        cargo = guilda_discord.get_role(id_cargo)
        if cargo:
            cargos_gerenciados.append(cargo)

    for membro in guilda_discord.members:
        if membro.bot:
            continue

        nomes_imunes = ["lider", "DIE HARD", "recrutador", "moderador", "caller", "SUB-LIDER"]
        ids_imunes = [CARGOS.get(nome) for nome in nomes_imunes if CARGOS.get(nome)]

        if any(c.id in ids_imunes for c in membro.roles):
            continue

        cargos_do_membro = [c for c in cargos_gerenciados if c in membro.roles]

        if not cargos_do_membro:
            continue

        nick = membro.display_name
        if " " in nick:
            nick = nick.split(" ", 1)[1] 

        async with aiohttp.ClientSession() as session:
            try:
                async with session.get(f"https://gameinfo.albiononline.com/api/gameinfo/search?q={nick}") as resp:
                    if resp.status != 200:
                        await asyncio.sleep(2) 
                        continue
                    
                    dados = await resp.json()
                    jogadores = dados.get('players', [])
                    jogador_encontrado = next((p for p in jogadores if p['Name'].lower() == nick.lower()), None)

                    rebaixar = False

                    if not jogador_encontrado:
                        rebaixar = True
                    else:
                        guild_id_jogador = jogador_encontrado.get('GuildId')
                        alliance_id_jogador = jogador_encontrado.get('AllianceId')

                        is_guilda = (guild_id_jogador == GUILDA_ALBION_ID)
                        is_alianca = (ALIANCA_ALBION_ID and alliance_id_jogador == ALIANCA_ALBION_ID)

                        if not is_guilda and not is_alianca:
                            rebaixar = True

                    if rebaixar:
                        await membro.remove_roles(*cargos_do_membro)
                        print(f"⚠️ {membro.display_name} foi rebaixado.")
                        
                        try:
                            await membro.send("⚠️ **Aviso Automático:** Seus cargos no Discord da guilda foram removidos porque nosso sistema detectou que você não está mais na guilda/aliança no jogo. Se isso for um erro, use o comando `!registrar` novamente!")
                        except discord.Forbidden:
                            pass 

            except Exception as e:
                print(f"Erro na auditoria do jogador {nick}: {e}")

        await asyncio.sleep(2)
        
    print("Base concluída.")

# ==========================================
# SISTEMA DE FARM DE PONTOS EM CALL (MONGODB)
# ==========================================

@tasks.loop(minutes=10)
async def farm_de_pontos():
    if not MONGO_URI:
        return

    for guilda in bot.guilds:
        cargos_bonus = [CARGOS.get("caller"), CARGOS.get("caller novato")]

        for canal_voz in guilda.voice_channels:
            membros_na_call = [m for m in canal_voz.members if not m.bot]

            if len(membros_na_call) >= MINIMO_PESSOAS_CALL:
                for membro in membros_na_call:
                    id_str = str(membro.id)
                    eh_caller = any(c.id in cargos_bonus for c in membro.roles)
                    
                    pontos_ganhos = (PONTOS_POR_CICLO * MULTIPLICADOR_CALLER) if eh_caller else PONTOS_POR_CICLO

                    await colecao_pontos.update_one(
                        {"_id": id_str},
                        {"$inc": {"pontos": pontos_ganhos}},
                        upsert=True
                    )

# ==========================================
#  EVENTOS DO BOT
# ==========================================

@bot.event
async def on_ready():
    print(f'🔥 Sistema Mestre online! Operando como {bot.user}.')
    
    if not auditoria_guilda.is_running():
        auditoria_guilda.start()
        
    if not limpeza_automatica.is_running():
        limpeza_automatica.start()

    if not farm_de_pontos.is_running():
        farm_de_pontos.start()

# ==========================================
# EVENTOS DE CARGO POR REAÇÃO (REACTION ROLES)
# ==========================================

@bot.event
async def on_raw_reaction_add(payload):
    if payload.message_id != MENSAGEM_CLASSES_ID:
        return
    
    if payload.member.bot:
        return

    emoji = str(payload.emoji)
    id_cargo = REACOES_CLASSES.get(emoji)

    if id_cargo:
        guild = bot.get_guild(payload.guild_id)
        cargo = guild.get_role(id_cargo)
        if cargo:
            try:
                await payload.member.add_roles(cargo)
                print(f"✅ {payload.member.display_name} pegou a classe de {cargo.name}.")
            except discord.Forbidden:
                print("❌ Erro: O cargo do bot precisa estar acima do cargo da classe.")

@bot.event
async def on_raw_reaction_remove(payload):
    if payload.message_id != MENSAGEM_CLASSES_ID:
        return
    
    guild = bot.get_guild(payload.guild_id)
    if not guild:
        return
    
    membro = guild.get_member(payload.user_id)
    if not membro or membro.bot:
        return

    emoji = str(payload.emoji)
    id_cargo = REACOES_CLASSES.get(emoji)

    if id_cargo:
        cargo = guild.get_role(id_cargo)
        if cargo:
            try:
                await membro.remove_roles(cargo)
                print(f"🔴 {membro.display_name} removeu a classe de {cargo.name}.")
            except discord.Forbidden:
                pass        
        
# ==========================================
#  CRIAR CALL
# ==========================================

@bot.event
async def on_voice_state_update(member, before, after):
    if after.channel and after.channel.id in CANAIS_GERADORES_IDS:
        guilda = member.guild
        categoria = after.channel.category
        
        permissoes = {
            guilda.default_role: discord.PermissionOverwrite(view_channel=False), 
            member: discord.PermissionOverwrite(view_channel=True, connect=True) 
        }
        
        for nome, id_cargo in CARGOS.items():
            cargo_obj = guilda.get_role(id_cargo)
            if cargo_obj: 
                permissoes[cargo_obj] = discord.PermissionOverwrite(view_channel=True, connect=True)
        
        novo_canal = await guilda.create_voice_channel(
            name=f"🎮 {member.display_name}",
            category=categoria,
            overwrites=permissoes
        )
        await member.move_to(novo_canal)

    if before.channel and before.channel.name.startswith("🎮") and before.channel.id not in CANAIS_GERADORES_IDS:
        if len(before.channel.members) == 0:
            await before.channel.delete()

# ==========================================
#  COMANDOS GERAIS E DE AJUDA
# ==========================================

@bot.command(name="ajuda", aliases=["help", "comandos"])
async def ajuda(ctx):
    await ctx.message.delete()
    
    embed = discord.Embed(
        title="📚 Central de Comandos da Guilda",
        description="Aqui estão todos os comandos disponíveis no servidor. Escolha o que você precisa:",
        color=discord.Color.gold()
    )
    
    embed.add_field(
        name="👤 Comandos de Membros", 
        value=(
            "`!registrar SeuNick` ➔ Vincula sua conta do jogo e libera seu acesso.\n"
            "`!pontos` (ou `!saldo`) ➔ Mostra quantos pontos de atividade você tem.\n"
            "`!ping` ➔ Verifica se os sistemas estão online."
        ), 
        inline=False
    )
    
    embed.add_field(
        name="⚔️ Comandos de Formação (LFG)", 
        value=(
            "`!vaga Conteúdo / Classe:Qtd / Hora:Minuto`\n"
            "➔ *Cria um painel interativo de PT.*\n"
            "➔ *Exemplo:* `!vaga ZvZ / Tank:2 / Healer:2 / 20:30`"
        ), 
        inline=False
    )
    
    # Verifica permissões para mostrar a área VIP da ajuda
    ids_permitidos = [CARGOS.get(nome) for nome in CARGOS_PERMITIDOS_ADICIONAR if CARGOS.get(nome)]
    tem_permissao = any(cargo.id in ids_permitidos for cargo in ctx.author.roles)
    
    if ctx.author.guild_permissions.administrator or tem_permissao:
        embed.add_field(
            name="⚙️ Comandos de Liderança (Restrito)", 
            value=(
                "`!adicionarpontos @membro 10` ➔ Adiciona pontos na conta de alguém.\n"
                "`!removerpontos @membro 10` ➔ Remove pontos da conta de alguém.\n"
                "`!painel_registro` ➔ Gera o painel fixo de boas-vindas no chat atual."
            ), 
            inline=False
        )
        
    embed.set_thumbnail(url=ctx.guild.icon.url if ctx.guild.icon else None)
    embed.set_footer(text="Dúvidas? Procure um moderador ou oficial da guilda.")
    
    try:
        await ctx.author.send(embed=embed)
    except discord.Forbidden:
        await ctx.send(f"⚠️ {ctx.author.mention}, sua DM está trancada! Aqui estão os comandos:", embed=embed, delete_after=60)


@bot.command()
async def ping(ctx):
    await ctx.send("Pong! Todos os sistemas operacionais.")

@bot.command(name="painel_registro")
@commands.has_permissions(administrator=True) 
async def painel_registro(ctx):
    await ctx.message.delete()
    embed = discord.Embed(
        title="🛡️ PORTAL DE REGISTRO 🛡️",
        description=(
            "Bem-vindo ao nosso servidor!\n\n"
            "Para ter acesso aos canais de voz e texto, você precisa confirmar sua conta do **Albion Online**.\n\n"
            "👉 **Para se registrar, digite aqui no chat o comando:**\n"
            "`!registrar SeuNickDoJogo`\n\n"
            "*Exemplo: `!registrar Zezinho`*"
        ),
        color=discord.Color.blue()
    )
    embed.set_thumbnail(url=ctx.guild.icon.url if ctx.guild.icon else None)
    await ctx.send(embed=embed)

@bot.command(name="registrar")
async def registrar(ctx, *, nick: str = None):
    await ctx.message.delete()
    if not nick:
        return await ctx.send(f"⚠️ {ctx.author.mention}, use `!registrar SeuNick`.", delete_after=10)
        
    msg_aviso = await ctx.send(f"🔍 Buscando **{nick}** nos servidores do Albion, aguarde um momento...")
    
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(f"https://gameinfo.albiononline.com/api/gameinfo/search?q={nick}") as resp:
                if resp.status != 200:
                    return await msg_aviso.edit(content="❌ Erro ao conectar com o Albion. Tente novamente.")
                
                dados = await resp.json()
                jogadores = dados.get('players', [])
                jogador_encontrado = next((p for p in jogadores if p['Name'].lower() == nick.lower()), None)
                        
                if not jogador_encontrado:
                    return await msg_aviso.edit(content=f"❌ O jogador **{nick}** não foi encontrado!")

                guild_id_jogador = jogador_encontrado.get('GuildId')
                alliance_id_jogador = jogador_encontrado.get('AllianceId')
                nome_correto = jogador_encontrado['Name']

                cargo_dar = None
                nova_tag = ""
                mensagem_final = ""

                if guild_id_jogador == GUILDA_ALBION_ID:
                    cargo_dar = ctx.guild.get_role(CARGOS["membro"])
                    nova_tag = f"{TAG_GUILDA} {nome_correto}"
                    mensagem_final = f"✅ **Sucesso!** Bem-vindo à guilda, {ctx.author.mention}!"
                elif ALIANCA_ALBION_ID and alliance_id_jogador == ALIANCA_ALBION_ID:
                    cargo_dar = ctx.guild.get_role(CARGOS["aliado"])
                    nova_tag = f"{TAG_ALIANCA} {nome_correto}"
                    mensagem_final = f"🤝 **Sucesso!** Aliado reconhecido, {ctx.author.mention}!"
                else:
                    return await msg_aviso.edit(content=f"❌ Acesso Negado: Você não pertence à Guilda/Aliança.")

                if cargo_dar:
                    try:
                        await ctx.author.add_roles(cargo_dar)
                    except discord.Forbidden:
                        pass
                try:
                    await ctx.author.edit(nick=nova_tag[:32])
                except discord.Forbidden:
                    pass 

                await msg_aviso.edit(content=mensagem_final)

    except Exception as e:
        await msg_aviso.edit(content=f"⚠️ Ocorreu um erro interno: {e}")

# ==========================================
# SISTEMA DE VAGAS (LFG)
# ==========================================

class BotaoDinamico(discord.ui.Button):
    def __init__(self, classe_nome, view_pai):
        super().__init__(label=classe_nome, style=discord.ButtonStyle.secondary)
        self.classe_nome = classe_nome
        self.view_pai = view_pai

    async def callback(self, interaction: discord.Interaction):
        await self.view_pai.processar_clique(interaction, self.classe_nome)

class ModalPuxarMembro(discord.ui.Modal, title="👑 Puxar Membro para a PT"):
    jogador = discord.ui.TextInput(
        label="Nome, @Nick ou ID numérico",
        style=discord.TextStyle.short,
        placeholder="Ex: @[DH] Zezinho",
        required=True
    )
    classe = discord.ui.TextInput(
        label="Nome da Vaga (Exatamente como no painel)",
        style=discord.TextStyle.short,
        placeholder="Ex: Tank, Suporte, DPS",
        required=True
    )

    def __init__(self, view_pai):
        super().__init__()
        self.view_pai = view_pai

    async def on_submit(self, interaction: discord.Interaction):
        usuario_input = self.jogador.value.strip()
        classe_escolhida = self.classe.value.strip()
        
        if classe_escolhida not in self.view_pai.max_vagas:
            return await interaction.response.send_message(f"❌ A classe `{classe_escolhida}` não existe.", ephemeral=True)

        usuario_final = None
        
        if usuario_input.startswith("<@") and usuario_input.endswith(">"):
            usuario_final = usuario_input 
        elif usuario_input.isdigit():
            usuario_final = f"<@{usuario_input}>"        
        else:
            nome_busca = usuario_input.lstrip('@').strip().lower()
            
            for membro in interaction.guild.members:
                if membro.display_name.lower() == nome_busca or membro.name.lower() == nome_busca:
                    usuario_final = membro.mention
                    break
            
            if not usuario_final:
                return await interaction.response.send_message(
                    f"❌ Não encontrei ninguém com o nome `{usuario_input}` no servidor.\n"
                    "*Dica: Digite exatamente igual ao apelido do Discord ou cole o ID numérico.*", 
                    ephemeral=True
                )

        await self.view_pai.forcar_insercao(interaction, usuario_final, classe_escolhida)


class PainelVagas(discord.ui.View):
    def __init__(self, conteudo, definicao_vagas, autor_id, unix_timestamp=None):
        super().__init__(timeout=None)
        self.conteudo = conteudo
        self.max_vagas = definicao_vagas
        self.autor_id = autor_id 
        self.unix_timestamp = unix_timestamp 
        self.jogadores = {classe: [] for classe in definicao_vagas}
        self.fila_espera = {classe: [] for classe in definicao_vagas}

        for classe in definicao_vagas.keys():
            self.add_item(BotaoDinamico(classe, self))
            
        botao_sair = discord.ui.Button(label="Sair da Lista", style=discord.ButtonStyle.danger, emoji="❌")
        botao_sair.callback = self.sair_callback
        self.add_item(botao_sair)

    def gerar_embed(self):
        titulo_destaque = f"💥 {self.conteudo.upper()} 💥"
        
        status_texto = "🟢 Formando Grupo"
        if self.unix_timestamp:
            status_texto += f" | ⏱️ **Começa:** <t:{self.unix_timestamp}:R>"
        
        embed = discord.Embed(
            title=titulo_destaque, 
            description=f"**Líder da PT:** <@{self.autor_id}>\n**Status:** {status_texto}\n━━━━━━━━━━━━━━━━━━━━━━\n",
            color=discord.Color.brand_red()
        )
        
        for classe, vagas_totais in self.max_vagas.items():
            inscritos = self.jogadores[classe]
            reserva = self.fila_espera[classe]
            texto_jogadores = "\n".join(inscritos) if inscritos else "*Vazio*"
            
            if reserva:
                texto_reserva = "\n".join([f"⏳ *{r} (Fila)*" for r in reserva])
                texto_final = f"{texto_jogadores}\n\n**⏱️ Fila:**\n{texto_reserva}"
            else:
                texto_final = texto_jogadores

            embed.add_field(name=f"🛡️ {classe} ({len(inscritos)}/{vagas_totais})", value=texto_final, inline=True)
            
        embed.set_footer(text="Clique nos botões abaixo para entrar ou sair da fila.")
        return embed

    async def promover_da_fila(self, interaction: discord.Interaction, classe: str):
        if len(self.jogadores[classe]) < self.max_vagas[classe] and len(self.fila_espera[classe]) > 0:
            proximo_jogador = self.fila_espera[classe].pop(0) 
            self.jogadores[classe].append(proximo_jogador)
            await interaction.channel.send(f"🎉 {proximo_jogador}, uma vaga abriu e você assumiu como **{classe}**!")

    async def processar_clique(self, interaction: discord.Interaction, classe: str):
        usuario = interaction.user.mention
        classe_antiga = None
        
        for c in self.jogadores:
            if usuario in self.jogadores[c]: 
                self.jogadores[c].remove(usuario)
                classe_antiga = c 
            if usuario in self.fila_espera[c]: 
                self.fila_espera[c].remove(usuario)

        if len(self.jogadores[classe]) < self.max_vagas[classe]:
            self.jogadores[classe].append(usuario)
            await interaction.response.edit_message(embed=self.gerar_embed(), view=self)
        else:
            if usuario not in self.fila_espera[classe]:
                self.fila_espera[classe].append(usuario)
                await interaction.response.edit_message(embed=self.gerar_embed(), view=self)
                await interaction.followup.send(f"📋 Fila de Espera para {classe}!", ephemeral=True)

        if classe_antiga and classe_antiga != classe:
            await self.promover_da_fila(interaction, classe_antiga)
            await interaction.message.edit(embed=self.gerar_embed(), view=self)

    async def sair_callback(self, interaction: discord.Interaction):
        usuario = interaction.user.mention
        removido = False
        classe_abandonada = None
        
        for c in self.jogadores:
            if usuario in self.jogadores[c]:
                self.jogadores[c].remove(usuario)
                removido = True
                classe_abandonada = c
            if usuario in self.fila_espera[c]:
                self.fila_espera[c].remove(usuario)
                removido = True
                
        if removido:
            await interaction.response.edit_message(embed=self.gerar_embed(), view=self)
            if classe_abandonada:
                await self.promover_da_fila(interaction, classe_abandonada)
                await interaction.message.edit(embed=self.gerar_embed(), view=self)
        else:
            await interaction.response.send_message("Você não está inscrito em nenhuma vaga.", ephemeral=True)

    async def abrir_modal_lider(self, interaction: discord.Interaction):
        if interaction.user.id != self.autor_id:
            return await interaction.response.send_message("❌ Acesso Negado: Apenas o criador pode puxar membros!", ephemeral=True)
        await interaction.response.send_modal(ModalPuxarMembro(self))

    async def forcar_insercao(self, interaction: discord.Interaction, usuario: str, classe: str):
        for c in self.jogadores:
            if usuario in self.jogadores[c]: 
                self.jogadores[c].remove(usuario)
            if usuario in self.fila_espera[c]: 
                self.fila_espera[c].remove(usuario)

        if len(self.jogadores[classe]) < self.max_vagas[classe]:
            self.jogadores[classe].append(usuario)
            await interaction.response.edit_message(embed=self.gerar_embed(), view=self)
            await interaction.followup.send(f"✅ Membro forçado na vaga de **{classe}**!", ephemeral=True)
        else:
            self.fila_espera[classe].append(usuario)
            await interaction.response.edit_message(embed=self.gerar_embed(), view=self)
            await interaction.followup.send(f"⚠️ Fila de Espera de **{classe}**.", ephemeral=True)

@bot.command(name="vaga")
async def vaga(ctx, *, texto: str = None):
    await ctx.message.delete()
    
    # Se o usuário digitou apenas "!vaga" sem nada, mostramos o manual completo
    if texto is None:
        embed_ajuda = discord.Embed(
            title="🛠️ Manual de Criação de PT (!vaga)",
            description="Aprenda a montar o seu painel interativo de vagas passo a passo.",
            color=discord.Color.blue()
        )
        embed_ajuda.add_field(
            name="🧩 Entendendo as Peças do Comando:",
            value=(
                "**1. Título:** O nome do seu conteúdo (Ex: `Gank`).\n"
                "**2. Vagas:** O nome da função e a quantidade de jogadores separados por dois-pontos (Ex: `Tank:2`).\n"
                "**3. Horário:** A hora que a PT vai sair (Opcional, Ex: `20:30`).\n"
                "*(Importante: Separe cada uma dessas peças usando uma barra `/`)*"
            ),
            inline=False
        )
        embed_ajuda.add_field(
            name="🎯 Exemplo 1: PT Agendada (Com Horário)",
            value=(
                "*Você quer montar um Gank mais tarde, às 21:30, e precisa de 1 Tank, 2 Suportes e 5 DPS.* \n"
                "**Você digita exatamente assim:**\n"
                "`!vaga 🗡️ Gank na Red / Tank:1 / Suporte:2 / DPS:5 / 21:30`"
            ),
            inline=False
        )
        embed_ajuda.add_field(
            name="⚡ Exemplo 2: PT Imediata (Sem Horário)",
            value=(
                "*Você quer montar um grupo de Fama agora mesmo e precisa de 1 Healer e 3 DPS.*\n"
                "**Você digita exatamente assim:**\n"
                "`!vaga 📖 Fama Group / Healer:1 / DPS:3`"
            ),
            inline=False
        )
        embed_ajuda.add_field(
            name="⚠️ Atenção aos Detalhes:",
            value=(
                "• Nunca esqueça de colocar os dois-pontos `:` para a quantidade de vagas (Errado: `Tank 2`, Certo: `Tank:2`).\n"
                "• Os nomes das classes que você escrever aqui (ex: DPS, Curandeiro, Suporte) serão **exatamente** os nomes que vão aparecer nos botões em que os jogadores vão clicar."
            ),
            inline=False
        )
        return await ctx.send(embed=embed_ajuda, delete_after=60)
    
    try:
        partes = texto.split('/')
        conteudo = partes[0].strip()
        definicao_vagas = {}
        horario_evento = None
        
        for parte in partes[1:]:
            parte_limpa = parte.strip()
            if not parte_limpa:
                continue
                
            if ":" in parte_limpa and len(parte_limpa) <= 5:
                partes_tempo = parte_limpa.split(':')
                if len(partes_tempo) == 2:
                    h, m = partes_tempo
                    if h.isdigit() and m.isdigit() and len(h) <= 2 and len(m) == 2:
                        horario_evento = parte_limpa
                        continue 
                        
            if ":" not in parte_limpa:
                raise ValueError()
                
            nome_classe, qtd = parte_limpa.split(':', 1)
            if not qtd.strip().isdigit():
                raise ValueError()
                
            definicao_vagas[nome_classe.strip()] = int(qtd.strip())
            
        if not definicao_vagas:
            raise ValueError()
            
    except Exception as e:
        # Mensagem de erro simples, convidando o usuário a ler o manual se precisar
        mensagem_erro = "⚠️ **Formato incorreto!** Faltou alguma barra (`/`) ou dois-pontos (`:`). Digite apenas `!vaga` para ver o manual de instruções."
        return await ctx.send(mensagem_erro, delete_after=60)

    unix_timestamp = None
    if horario_evento:
        try:
            h_str, m_str = horario_evento.split(':')
            fuso = timezone(timedelta(hours=-4))
            agora = datetime.now(fuso)
            data_evento = agora.replace(hour=int(h_str), minute=int(m_str), second=0, microsecond=0)
            if data_evento < agora:
                data_evento += timedelta(days=1)
            unix_timestamp = int(data_evento.timestamp())
        except:
            pass

    painel = PainelVagas(conteudo, definicao_vagas, ctx.author.id, unix_timestamp)
    embed_inicial = painel.gerar_embed()
    
    id_cargo_membro = CARGOS.get("membro")
    mencao_cargo = f"<@&{id_cargo_membro}>" if id_cargo_membro else "@everyone"

    await ctx.send(content=f"📢 {mencao_cargo}", embed=embed_inicial, view=painel)
# ==========================================
#  COMANDOS DO SISTEMA DE PONTOS
# ==========================================

@bot.command(name="pontos", aliases=["saldo"])
async def ver_pontos(ctx, membro: discord.Member = None):
    membro_alvo = membro or ctx.author 
    
    if not MONGO_URI:
        return await ctx.send("⚠️ O Banco de Dados de pontos está inacessível no momento.")

    documento = await colecao_pontos.find_one({"_id": str(membro_alvo.id)})
    saldo = documento.get("pontos", 0) if documento else 0
    
    embed = discord.Embed(
        title="💰 Banco da Guilda - Saldo de Atividade",
        description=f"O jogador {membro_alvo.mention} possui **{saldo} pontos** acumulados.",
        color=discord.Color.green()
    )
    embed.set_thumbnail(url=membro_alvo.display_avatar.url)
    embed.set_footer(text="Participe de conteúdos em calls com 5+ membros para pontuar!")
    await ctx.send(embed=embed)

@bot.command(name="adicionarpontos")
async def adicionar_pontos_manual(ctx, membro: discord.Member, quantidade: int):
    # 1. Blindagem de Permissão
    ids_permitidos = [CARGOS.get(nome) for nome in CARGOS_PERMITIDOS_ADICIONAR if CARGOS.get(nome)]
    tem_permissao = any(cargo.id in ids_permitidos for cargo in ctx.author.roles)
    
    # Administradores do servidor (Dono) sempre conseguem usar, independentemente do cargo
    if not tem_permissao and not ctx.author.guild_permissions.administrator:
        return await ctx.send("❌ Acesso Negado: Seu cargo não tem permissão para adicionar pontos.")

    # 2. Verificações de Segurança
    if not MONGO_URI:
        return await ctx.send("⚠️ Banco de dados offline.")
        
    if quantidade <= 0:
        return await ctx.send("❌ A quantidade deve ser um número positivo.")
        
    if quantidade > LIMITE_ADICAO_POR_VEZ:
        return await ctx.send(f"❌ Medida de Segurança: Você só pode adicionar até **{LIMITE_ADICAO_POR_VEZ} pontos** por vez.")

    # 3. Execução
    await colecao_pontos.update_one(
        {"_id": str(membro.id)},
        {"$inc": {"pontos": quantidade}},
        upsert=True
    )
    await ctx.send(f"✅ Foram adicionados **{quantidade} pontos** na carteira de {membro.mention} com sucesso!")


@bot.command(name="removerpontos")
async def remover_pontos_manual(ctx, membro: discord.Member, quantidade: int):
    # 1. Blindagem de Permissão (Usa a lista mais rigorosa)
    ids_permitidos = [CARGOS.get(nome) for nome in CARGOS_PERMITIDOS_REMOVER if CARGOS.get(nome)]
    tem_permissao = any(cargo.id in ids_permitidos for cargo in ctx.author.roles)
    
    if not tem_permissao and not ctx.author.guild_permissions.administrator:
        return await ctx.send("❌ Acesso Restrito: Apenas a Alta Cúpula pode remover pontos dos membros.")

    # 2. Verificações de Segurança
    if not MONGO_URI:
        return await ctx.send("⚠️ Banco de dados offline.")
        
    if quantidade <= 0:
        return await ctx.send("❌ A quantidade deve ser um número positivo.")

    # 3. Checagem de Saldo (Impede carteira negativa)
    documento = await colecao_pontos.find_one({"_id": str(membro.id)})
    saldo_atual = documento.get("pontos", 0) if documento else 0

    if saldo_atual < quantidade:
        return await ctx.send(f"❌ Operação cancelada. O membro possui apenas **{saldo_atual} pontos**.")

    # 4. Execução
    await colecao_pontos.update_one(
        {"_id": str(membro.id)},
        {"$inc": {"pontos": -quantidade}},
        upsert=True
    )
    await ctx.send(f"🛑 Foram removidos **{quantidade} pontos** da carteira de {membro.mention}.")

    @bot.command(name="relatorio")
    async def gerar_relatorio_pontos(ctx):
        # 1. Blindagem de Permissão (Apenas Líder e Sub-Líder)
        ids_permitidos = [CARGOS.get(nome) for nome in CARGOS_PERMITIDOS_REMOVER if CARGOS.get(nome)]
        tem_permissao = any(cargo.id in ids_permitidos for cargo in ctx.author.roles)
        
        if not tem_permissao and not ctx.author.guild_permissions.administrator:
            return await ctx.send("❌ Acesso Restrito: Apenas a Alta Cúpula pode gerar o relatório geral de atividade.")

        if not MONGO_URI:
            return await ctx.send("⚠️ Banco de dados offline.")

        mensagem_aviso = await ctx.send("🔄 **Extraindo dados da nuvem e montando a planilha...** Aguarde um instante.")

        try:
            # Puxa tudo o que existe no MongoDB
            cursor = colecao_pontos.find({})
            documentos = await cursor.to_list(length=None)

            if not documentos:
                return await mensagem_aviso.edit(content="❌ O banco de dados está vazio no momento.")

            # Cria a planilha em uma "nuvem de memória" (não precisa de HD para salvar)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Relatório da Guilda"

            # Cores e Estilos para a tabela ficar com cara corporativa/profissional
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            zebra_fill = PatternFill(start_color="F2F6FA", end_color="F2F6FA", fill_type="solid")
            font_header = Font(color="FFFFFF", bold=True)
            align_center = Alignment(horizontal="center", vertical="center")

            # Monta os títulos das colunas
            headers = ["ID do Jogador", "Nick na Guilda (Discord)", "Pontos Acumulados"]
            ws.append(headers)

            # Pinta o cabeçalho
            for col_num, _ in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = font_header
                cell.fill = header_fill
                cell.alignment = align_center

            # Organiza os dados e preenche a planilha
            linha_atual = 2
            for doc in documentos:
                id_discord = int(doc["_id"])
                pontos = doc.get("pontos", 0)

                # Tenta achar o membro no servidor para colocar o Nick com a TAG certinha
                membro = ctx.guild.get_member(id_discord)
                nome_jogador = membro.display_name if membro else "Membro Desligado (Fora do Discord)"

                ws.append([str(id_discord), nome_jogador, pontos])

                # Pinta uma linha sim, uma não (Zebrado) para facilitar a leitura do líder
                if linha_atual % 2 == 0:
                    for col_idx in range(1, 4):
                        ws.cell(row=linha_atual, column=col_idx).fill = zebra_fill
                
                linha_atual += 1

            # Arruma a largura das colunas
            ws.column_dimensions["A"].width = 25
            ws.column_dimensions["B"].width = 40
            ws.column_dimensions["C"].width = 20

            # Congela o cabeçalho para quando rolar a tela pra baixo no Excel
            ws.freeze_panes = "A2"

            # Empacota o arquivo na memória para envio
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            # Monta o nome do arquivo com a data de hoje (ex: Relatorio_Pontos_26_06_2026.xlsx)
            nome_arquivo = f"Relatorio_Pontos_{datetime.now().strftime('%d_%m_%Y')}.xlsx"
            arquivo_discord = discord.File(fp=buffer, filename=nome_arquivo)
            
            # Envia o arquivo no chat e apaga o aviso
            await mensagem_aviso.delete()
            await ctx.send(content="📊 **Relatório Extraído com Sucesso!**\nAqui está a planilha oficial de atividade da guilda atualizada agora:", file=arquivo_discord)

        except Exception as e:
            await mensagem_aviso.edit(content=f"⚠️ Erro ao gerar a planilha: {e}")

# ==========================================
#  INICIALIZAÇÃO
# ==========================================
keep_alive() 
bot.run(os.getenv('TOKEN_DO_BOT'))